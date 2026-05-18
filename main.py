import os
import re
import pandas as pd
import docx
from deep_translator import GoogleTranslator
from openpyxl.styles import Font, PatternFill, Alignment


try:
    import win32com.client as win32
except ImportError:
    win32 = None
    print("Peringatan: Library 'pywin32' belum terinstall. Fitur konversi otomatis .doc ke .docx dimatikan.")
    print("Silakan install melalui CMD dengan perintah: pip install pywin32\n")

def bersihkan_teks(teks):
    if not teks: return ""
    teks = str(teks).strip()
    teks = re.sub(r'[\t\uf0b7]', ' ', teks)
    teks = re.sub(r' +', ' ', teks)
    return teks.strip()

def teks_kunci(teks):
    return re.sub(r'[^a-zA-Z0-9]', '', str(teks).lower())

def apakah_baris_sampah(teks_materi):
    teks_uji = teks_kunci(teks_materi)
    sampah = ["cpl", "cpmk", "rubrik", "naskahpresentasi", "materipembelajaran"]
    
    if len(teks_uji) < 3: return True 
    for s in sampah:
        if s == teks_uji or (s in teks_uji and len(teks_materi) < 25): 
            return True
    return False

def konversi_doc_ke_docx(folder_path):
    """Fungsi untuk mencari file .doc dan mengonversinya menjadi .docx"""
    if win32 is None:
        return # Keluar jika pywin32 tidak terinstall

    doc_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".doc") and not f.startswith("~")]
    
    if not doc_files:
        return # Jika tidak ada file .doc, langsung lanjut ke proses ekstraksi
        
    print(f"🔄 Mendeteksi {len(doc_files)} file berformat .doc lama. Memulai konversi ke .docx...")
    
    try:
        # Membuka Microsoft Word di latar belakang
        word = win32.Dispatch('Word.Application')
        word.Visible = False
        
        for filename in doc_files:
            in_file = os.path.abspath(os.path.join(folder_path, filename))
            out_filename = filename + "x" 
            out_file = os.path.abspath(os.path.join(folder_path, out_filename))
            

            if not os.path.exists(out_file):
                print(f"   -> Mengonversi: {filename}")
                try:
                    doc = word.Documents.Open(in_file)
                    doc.SaveAs(out_file, FileFormat=16)
                    doc.Close()
                except Exception as e:
                    print(f"Gagal mengonversi {filename}: {e}")
                    
        word.Quit()
        print("Konversi selesai.\n")
        
    except Exception as e:
        print(f"Gagal memanggil sistem Microsoft Word: {e}\n")


def ekstrak_materi_rps(folder_path):
    # JALANKAN PROSES KONVERSI TERLEBIH DAHULU (Jika ada)
    konversi_doc_ke_docx(folder_path)
    
    print("MEMULAI EKSTRAKSI DOKUMEN...")
    
    all_data = []
    translator = GoogleTranslator(source='id', target='en')
    
    # Ambil semua file yang sekarang berformat .docx
    files = [f for f in os.listdir(folder_path) if f.endswith(".docx") and not f.startswith("~")]
    total_files = len(files)

    for idx, filename in enumerate(files, 1):
        file_path = os.path.join(folder_path, filename)
        print(f"[{idx}/{total_files}] Membaca: {filename[:25]}... ", end="")
        
        try:
            doc = docx.Document(file_path)
            nama_mk = filename.replace(".docx", "").strip()

            tabel_utama_ditemukan = False
            col_materi = -1
            col_waktu = -1
            materi_per_minggu = {}
            minggu_saat_ini = "1"

            for table in doc.tables:
                if len(table.rows) == 0: continue
                baris_mulai = -1
                
                # CARI HEADER DENGAN KUNCI GANDA
                for row_idx, row in enumerate(table.rows):
                    seluruh_sel = [teks_kunci(c.text) for c in row.cells]
                    
                    ada_materi = False
                    ada_waktu = False
                    ada_kemampuan = False
                    tmp_col_materi = -1
                    tmp_col_waktu = -1
                    
                    for i, sel in enumerate(seluruh_sel):
                        if any(k in sel for k in ["materi", "bahankajian", "pokokbahasan"]):
                            ada_materi = True
                            tmp_col_materi = i
                        if any(k in sel for k in ["minggu", "pertemuan", "tatapmuka", "waktu", "no"]):
                            ada_waktu = True
                            tmp_col_waktu = i
                        if any(k in sel for k in ["kemampuan", "cpmk", "indikator"]):
                            ada_kemampuan = True
                            
                    if ada_materi and ada_waktu and ada_kemampuan:
                        col_materi = tmp_col_materi
                        col_waktu = tmp_col_waktu
                        baris_mulai = row_idx + 1
                        tabel_utama_ditemukan = True
                        break 

                # Jika tidak ada header, tapi sebelumnya tabel sudah ketemu -> Ini Split Table
                if baris_mulai == -1 and tabel_utama_ditemukan:
                    teks_awal = teks_kunci(table.rows[0].cells[0].text)
                    
                    if any(k in teks_awal for k in ["aturan", "kehadiran", "evaluasi"]):
                        tabel_utama_ditemukan = False
                        continue
                    baris_mulai = 0 

                # AMBIL DATA MATERI
                if tabel_utama_ditemukan and baris_mulai != -1:
                    for row in table.rows[baris_mulai:]:
                        if len(row.cells) <= max(col_waktu, col_materi): continue
                        
                        teks_baris_full = teks_kunci(" ".join([c.text for c in row.cells]))

                        # berhenti jika ketemu kata kunci referensi
                        if any(k in teks_baris_full for k in ["daftarreferensi", "daftarpustaka"]):
                            tabel_utama_ditemukan = False
                            break 

                        teks_waktu = bersihkan_teks(row.cells[col_waktu].text)
                        teks_materi = bersihkan_teks(row.cells[col_materi].text)

                        # VALIDASI ANGKA: Kolom waktu wajib punya angka
                        if teks_waktu and len(teks_waktu) <= 15 and any(char.isdigit() for char in teks_waktu):
                            minggu_saat_ini = teks_waktu
                        
                        if minggu_saat_ini not in materi_per_minggu:
                            materi_per_minggu[minggu_saat_ini] = []

                        if any(k in teks_baris_full for k in ["tengahsemester", "akhirsemester", "uts", "uas"]):
                            for c in row.cells:
                                c_bersih = bersihkan_teks(c.text)
                                if "tengah semester" in c_bersih.lower() or "akhir semester" in c_bersih.lower():
                                    if not materi_per_minggu[minggu_saat_ini] or materi_per_minggu[minggu_saat_ini][-1] != c_bersih:
                                        materi_per_minggu[minggu_saat_ini].append(c_bersih)
                                    break 
                            continue 


                        if teks_materi and not apakah_baris_sampah(teks_materi):
                            if not materi_per_minggu[minggu_saat_ini] or materi_per_minggu[minggu_saat_ini][-1] != teks_materi:
                                materi_per_minggu[minggu_saat_ini].append(teks_materi)
            
            if not materi_per_minggu:
                print("GAGAL (Tabel materi utama tidak terdeteksi)")
                continue

            # GABUNGKAN PER PERTEMUAN DAN TERJEMAHKAN
            final_materi_list = []
            pertemuan_labels = []
            
            for minggu, list_materi in materi_per_minggu.items():
                if list_materi:
                    final_materi_list.append("\n".join(list_materi))
                    pertemuan_labels.append(minggu)

            materi_inggris_list = []
            try:
                # memisahkan baris dengan HTML agar Google Translator tidak menghapus baris baru
                teks_terjemahan = [m.replace('\n', ' <br> ') for m in final_materi_list]
                hasil_inggris = translator.translate_batch(teks_terjemahan)
                materi_inggris_list = [h.replace(' <br> ', '\n').replace('<br>', '\n') for h in hasil_inggris]
            except Exception:
                for m in final_materi_list:
                    try:
                        res = translator.translate(m.replace('\n', ' <br> '))
                        materi_inggris_list.append(res.replace(' <br> ', '\n').replace('<br>', '\n'))
                    except:
                        materi_inggris_list.append("")

            # SUSUN BARIS EXCEL
            for i in range(len(final_materi_list)):
                all_data.append({
                    "Kode Mata Kuliah": nama_mk,
                    "Pertemuan": pertemuan_labels[i],
                    "Materi Indonesia": final_materi_list[i],
                    "Materi Inggris": materi_inggris_list[i] if i < len(materi_inggris_list) else ""
                })
            
            for _ in range(3):
                all_data.append({"Kode Mata Kuliah": "", "Pertemuan": "", "Materi Indonesia": "", "Materi Inggris": ""})
                
            print(f" Selesai ({len(final_materi_list)} Pertemuan Berhasil Diekstrak)")
            
        except Exception as e:
            print(f"ERROR: {e}")

    # PEMBUATAN EXCEL
    if all_data:
        print("\nMenyusun dan merapikan file Excel...")
        df = pd.DataFrame(all_data)
        output_file = os.path.join(folder_path, "Extrak_output.xlsx")
        
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Data_RPS')
                worksheet = writer.sheets['Data_RPS']
                
                header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                header_align = Alignment(horizontal="center", vertical="center")
                
                for col in range(1, 5):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                
                worksheet.column_dimensions['A'].width = 30 
                worksheet.column_dimensions['B'].width = 15 
                worksheet.column_dimensions['C'].width = 75 
                worksheet.column_dimensions['D'].width = 75 
                
                wrap_alignment = Alignment(wrap_text=True, vertical="top")
                for row in worksheet.iter_rows(min_row=2, max_row=len(all_data)+1, min_col=1, max_col=4):
                    for cell in row:
                        cell.alignment = wrap_alignment

            print(f"File Excel selesai dibuat, buka di:\n {output_file}")
            
        except PermissionError:
            print("\n GAGAL MENYIMPAN: File Excel sedang terbuka! Tutup dulu file Excel-nya lalu jalankan ulang.")
    else:
        print("\n Tidak ada data yang berhasil diekstrak.")

LOKASI_FOLDER = r"C:\Users\Sofiandi\Downloads\RPS_FOLDER" 
ekstrak_materi_rps(LOKASI_FOLDER)
